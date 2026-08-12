#!/usr/bin/env python3

import csv
import io
import os
import re
import threading
import urllib.error
import urllib.parse
import urllib.request
import tkinter as tk

from tkinter import ttk, filedialog, messagebox


# ----------------------------------------------------------------
# Constants
# ----------------------------------------------------------------

# Excel maximum rows per worksheet
EXCEL_ROW_LIMIT = 1_048_576

# Excel maximum length of a sheet name
SHEET_NAME_LIMIT = 31

# Characters Excel does not allow in a sheet name
INVALID_SHEET_CHARS = "[]:*?/\\"

# Optional comment used to name the sheet of a query, e.g.
#     -- tab: Daily totals
TAB_NAME_PATTERN = re.compile(
    r"^\s*(?:--|#)\s*(?:tab|sheet|name|nome)\s*[:=]\s*(.+?)\s*$",
    re.IGNORECASE
)


# ----------------------------------------------------------------
# SQL splitting
# ----------------------------------------------------------------

def split_statements(text):
    """
    Split a script into single statements on ';'.

    Semicolons inside string literals, quoted identifiers,
    line comments and block comments are ignored.

    Fragments that contain only comments or whitespace
    are dropped.
    """

    statements = []

    current = []

    has_code = False

    in_single = False
    in_double = False
    in_backtick = False
    in_line_comment = False
    in_block_comment = False

    index = 0

    length = len(text)

    while index < length:

        char = text[index]

        following = (
            text[index + 1]
            if index + 1 < length
            else ""
        )

        # ----- inside a line comment -----

        if in_line_comment:

            current.append(char)

            if char == "\n":
                in_line_comment = False

            index += 1

            continue

        # ----- inside a block comment -----

        if in_block_comment:

            if char == "*" and following == "/":

                current.append(char)
                current.append(following)

                in_block_comment = False

                index += 2

                continue

            current.append(char)

            index += 1

            continue

        # ----- inside a quoted section -----

        if in_single or in_double or in_backtick:

            current.append(char)

            if char == "\\" and following:

                current.append(following)

                index += 2

                continue

            if in_single and char == "'":
                in_single = False

            elif in_double and char == '"':
                in_double = False

            elif in_backtick and char == "`":
                in_backtick = False

            index += 1

            continue

        # ----- normal code -----

        if char == "-" and following == "-":

            in_line_comment = True

            current.append(char)
            current.append(following)

            index += 2

            continue

        if char == "#":

            in_line_comment = True

            current.append(char)

            index += 1

            continue

        if char == "/" and following == "*":

            in_block_comment = True

            current.append(char)
            current.append(following)

            index += 2

            continue

        if char == ";":

            if has_code:
                statements.append("".join(current))

            current = []

            has_code = False

            index += 1

            continue

        if char == "'":
            in_single = True

        elif char == '"':
            in_double = True

        elif char == "`":
            in_backtick = True

        if not char.isspace():
            has_code = True

        current.append(char)

        index += 1

    if has_code:
        statements.append("".join(current))

    return [
        statement.strip()
        for statement in statements
        if statement.strip()
    ]


def statement_name(statement, position):
    """
    Return the sheet name of a statement.

    Uses the first '-- tab: <name>' comment found,
    otherwise falls back to 'Query <position>'.
    """

    for line in statement.splitlines():

        match = TAB_NAME_PATTERN.match(line)

        if match:

            name = " ".join(match.group(1).split())

            if name:
                return name

    return f"Query {position}"


def clean_sheet_name(name, used_names):
    """
    Make a name usable as an Excel sheet name and
    unique inside the workbook.
    """

    cleaned = "".join(
        " " if char in INVALID_SHEET_CHARS else char
        for char in name
    )

    cleaned = " ".join(cleaned.split())

    # Excel rejects a leading or trailing apostrophe
    cleaned = cleaned.strip("'")

    if not cleaned:
        cleaned = "Query"

    cleaned = cleaned[:SHEET_NAME_LIMIT]

    base = cleaned

    counter = 2

    while cleaned.lower() in used_names:

        suffix = f" ({counter})"

        cleaned = (
            base[:SHEET_NAME_LIMIT - len(suffix)]
            + suffix
        )

        counter += 1

    used_names.add(cleaned.lower())

    return cleaned


def format_rows(count):
    """
    Return a readable row count, e.g. '1 row', '12,345 rows'.
    """

    if count == 1:
        return "1 row"

    return f"{count:,} rows"


def clean_file_name(name):
    """
    Make a name usable inside a file name.
    """

    cleaned = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        name
    ).strip("._-")

    return cleaned or "query"


class ClickHouseExporter:
    def __init__(self, root):
        self.root = root
        self.root.title("ClickHouse Query Exporter")
        self.root.geometry("950x720")
        self.root.minsize(800, 600)

        self.create_ui()

    def create_ui(self):

        # ------------------------------------------------------------
        # Connection
        # ------------------------------------------------------------

        connection_frame = ttk.LabelFrame(
            self.root,
            text="ClickHouse Connection"
        )

        connection_frame.pack(
            fill="x",
            padx=10,
            pady=10
        )

        ttk.Label(
            connection_frame,
            text="Host:"
        ).grid(
            row=0,
            column=0,
            padx=5,
            pady=5,
            sticky="w"
        )

        self.host_var = tk.StringVar(value="localhost")

        ttk.Entry(
            connection_frame,
            textvariable=self.host_var,
            width=25
        ).grid(
            row=0,
            column=1,
            padx=5,
            pady=5
        )

        ttk.Label(
            connection_frame,
            text="Port:"
        ).grid(
            row=0,
            column=2,
            padx=5,
            pady=5
        )

        self.port_var = tk.StringVar(value="8123")

        ttk.Entry(
            connection_frame,
            textvariable=self.port_var,
            width=8
        ).grid(
            row=0,
            column=3,
            padx=5,
            pady=5
        )

        ttk.Label(
            connection_frame,
            text="Database:"
        ).grid(
            row=0,
            column=4,
            padx=5,
            pady=5
        )

        self.database_var = tk.StringVar(value="default")

        ttk.Entry(
            connection_frame,
            textvariable=self.database_var,
            width=18
        ).grid(
            row=0,
            column=5,
            padx=5,
            pady=5
        )

        ttk.Label(
            connection_frame,
            text="User:"
        ).grid(
            row=1,
            column=0,
            padx=5,
            pady=5,
            sticky="w"
        )

        self.user_var = tk.StringVar(value="default")

        ttk.Entry(
            connection_frame,
            textvariable=self.user_var,
            width=25
        ).grid(
            row=1,
            column=1,
            padx=5,
            pady=5
        )

        ttk.Label(
            connection_frame,
            text="Password:"
        ).grid(
            row=1,
            column=2,
            padx=5,
            pady=5
        )

        self.password_var = tk.StringVar()

        self.password_entry = ttk.Entry(
            connection_frame,
            textvariable=self.password_var,
            show="*",
            width=25
        )

        self.password_entry.grid(
            row=1,
            column=3,
            columnspan=2,
            padx=5,
            pady=5,
            sticky="w"
        )

        self.show_password_var = tk.BooleanVar()

        ttk.Checkbutton(
            connection_frame,
            text="Show",
            variable=self.show_password_var,
            command=self.toggle_password
        ).grid(
            row=1,
            column=5,
            padx=5
        )

        self.test_button = ttk.Button(
            connection_frame,
            text="Test Connection",
            command=self.test_connection
        )

        self.test_button.grid(
            row=2,
            column=0,
            columnspan=2,
            padx=5,
            pady=8,
            sticky="w"
        )

        # ------------------------------------------------------------
        # Query
        # ------------------------------------------------------------

        query_frame = ttk.LabelFrame(
            self.root,
            text="SQL Queries (separate them with ;)"
        )

        query_frame.pack(
            fill="both",
            expand=True,
            padx=10,
            pady=5
        )

        self.query_text = tk.Text(
            query_frame,
            wrap="none",
            font=("Consolas", 10)
        )

        query_scroll_y = ttk.Scrollbar(
            query_frame,
            orient="vertical",
            command=self.query_text.yview
        )

        query_scroll_x = ttk.Scrollbar(
            query_frame,
            orient="horizontal",
            command=self.query_text.xview
        )

        self.query_text.configure(
            yscrollcommand=query_scroll_y.set,
            xscrollcommand=query_scroll_x.set
        )

        self.query_text.grid(
            row=0,
            column=0,
            sticky="nsew"
        )

        query_scroll_y.grid(
            row=0,
            column=1,
            sticky="ns"
        )

        query_scroll_x.grid(
            row=1,
            column=0,
            sticky="ew"
        )

        ttk.Label(
            query_frame,
            text=(
                "Each query is written to its own Excel sheet. "
                "Name a sheet by putting  -- tab: My name  "
                "above the query."
            )
        ).grid(
            row=2,
            column=0,
            columnspan=2,
            padx=5,
            pady=(3, 5),
            sticky="w"
        )

        query_frame.rowconfigure(0, weight=1)
        query_frame.columnconfigure(0, weight=1)

        self.query_text.insert(
            "1.0",
            "-- tab: Server\n"
            "SELECT\n"
            "    now() AS current_time,\n"
            "    version() AS server_version;\n"
            "\n"
            "-- tab: Databases\n"
            "SELECT\n"
            "    name,\n"
            "    engine\n"
            "FROM system.databases\n"
            "ORDER BY name"
        )

        # ------------------------------------------------------------
        # Output
        # ------------------------------------------------------------

        output_frame = ttk.LabelFrame(
            self.root,
            text="Output"
        )

        output_frame.pack(
            fill="x",
            padx=10,
            pady=10
        )

        ttk.Label(
            output_frame,
            text="Format:"
        ).grid(
            row=0,
            column=0,
            padx=5,
            pady=5
        )

        self.format_var = tk.StringVar(value="Excel")

        self.format_combo = ttk.Combobox(
            output_frame,
            textvariable=self.format_var,
            values=["CSV", "Excel"],
            state="readonly",
            width=10
        )

        self.format_combo.grid(
            row=0,
            column=1,
            padx=5,
            pady=5
        )

        self.export_button = ttk.Button(
            output_frame,
            text="Run Queries and Save",
            command=self.start_export
        )

        self.export_button.grid(
            row=0,
            column=2,
            padx=15,
            pady=5
        )

        ttk.Label(
            output_frame,
            text=(
                "Excel: one file, one sheet per query. "
                "CSV: one file per query."
            )
        ).grid(
            row=0,
            column=3,
            padx=5,
            pady=5,
            sticky="w"
        )

        # ------------------------------------------------------------
        # Progress
        # ------------------------------------------------------------

        self.progress = ttk.Progressbar(
            self.root,
            mode="indeterminate"
        )

        self.progress.pack(
            fill="x",
            padx=10,
            pady=(0, 5)
        )

        self.status_var = tk.StringVar(
            value="Ready"
        )

        self.status_label = ttk.Label(
            self.root,
            textvariable=self.status_var
        )

        self.status_label.pack(
            fill="x",
            padx=10,
            pady=(0, 10)
        )

    # ------------------------------------------------------------
    # Password
    # ------------------------------------------------------------

    def toggle_password(self):

        if self.show_password_var.get():
            self.password_entry.config(show="")
        else:
            self.password_entry.config(show="*")

    # ------------------------------------------------------------
    # Status
    # ------------------------------------------------------------

    def set_status(self, text):

        self.root.after(
            0,
            lambda value=text:
            self.status_var.set(value)
        )

    # ------------------------------------------------------------
    # HTTP connection
    # ------------------------------------------------------------

    def execute_query(
        self,
        query,
        output_format="CSVWithNames"
    ):

        host = self.host_var.get().strip()
        port = self.port_var.get().strip()
        database = self.database_var.get().strip()
        user = self.user_var.get().strip()
        password = self.password_var.get()

        params = {
            "database": database,
            "default_format": output_format
        }

        url = (
            f"http://{host}:{port}/?"
            + urllib.parse.urlencode(params)
        )

        request = urllib.request.Request(
            url,
            data=query.encode("utf-8"),
            method="POST"
        )

        if user:
            request.add_header(
                "X-ClickHouse-User",
                user
            )

        if password:
            request.add_header(
                "X-ClickHouse-Key",
                password
            )

        try:
            return urllib.request.urlopen(
                request,
                timeout=60
            )

        except urllib.error.HTTPError as e:

            error = e.read().decode(
                "utf-8",
                errors="replace"
            )

            raise Exception(error)

        except urllib.error.URLError as e:
            raise Exception(
                f"Connection error:\n{e}"
            )

    # ------------------------------------------------------------
    # Test connection
    # ------------------------------------------------------------

    def test_connection(self):

        self.status_var.set(
            "Testing connection..."
        )

        self.test_button.config(
            state="disabled"
        )

        self.progress.start()

        threading.Thread(
            target=self.test_connection_thread,
            daemon=True
        ).start()

    def test_connection_thread(self):

        try:

            response = self.execute_query(
                "SELECT version()"
            )

            result = response.read().decode(
                "utf-8"
            ).strip()

            # CSVWithNames returns:
            # version()
            # 26.x.x

            lines = result.splitlines()

            version = (
                lines[-1]
                if lines
                else "Unknown"
            )

            self.root.after(
                0,
                lambda: self.connection_success(version)
            )

        except Exception as e:

            self.root.after(
                0,
                lambda error=str(e):
                self.connection_error(error)
            )

    def connection_success(self, version):

        self.progress.stop()

        self.test_button.config(
            state="normal"
        )

        self.status_var.set(
            f"Connected - ClickHouse {version}"
        )

        messagebox.showinfo(
            "Connection successful",
            f"Connected successfully.\n\n"
            f"ClickHouse version:\n{version}"
        )

    def connection_error(self, error):

        self.progress.stop()

        self.test_button.config(
            state="normal"
        )

        self.status_var.set(
            "Connection failed"
        )

        messagebox.showerror(
            "Connection error",
            error
        )

    # ------------------------------------------------------------
    # Export
    # ------------------------------------------------------------

    def start_export(self):

        script = self.query_text.get(
            "1.0",
            "end"
        ).strip()

        queries = split_statements(script)

        if not queries:

            messagebox.showwarning(
                "Query missing",
                "Insert at least one SQL query."
            )

            return

        names = [
            statement_name(query, position)
            for position, query in enumerate(
                queries,
                start=1
            )
        ]

        output_format = self.format_var.get()

        if output_format == "CSV":

            title = (
                "Save CSV"
                if len(queries) == 1
                else f"Save CSV - {len(queries)} files "
                     f"will be created"
            )

            filename = filedialog.asksaveasfilename(
                title=title,
                defaultextension=".csv",
                filetypes=[
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ]
            )

        else:

            filename = filedialog.asksaveasfilename(
                title=(
                    f"Save Excel - {len(queries)} "
                    f"sheet(s)"
                ),
                defaultextension=".xlsx",
                filetypes=[
                    ("Excel files", "*.xlsx"),
                    ("All files", "*.*")
                ]
            )

        if not filename:
            return

        self.export_button.config(
            state="disabled"
        )

        self.test_button.config(
            state="disabled"
        )

        self.progress.start()

        self.status_var.set(
            f"Executing {len(queries)} query(ies)..."
        )

        threading.Thread(
            target=self.export_thread,
            args=(
                queries,
                names,
                filename,
                output_format
            ),
            daemon=True
        ).start()

    def export_thread(
        self,
        queries,
        names,
        filename,
        output_format
    ):

        try:

            if output_format == "CSV":

                completed, error = self.export_csv(
                    queries,
                    names,
                    filename
                )

            else:

                completed, error = self.export_excel(
                    queries,
                    names,
                    filename
                )

            self.root.after(
                0,
                lambda: self.export_success(
                    filename,
                    completed,
                    error
                )
            )

        except Exception as e:

            self.root.after(
                0,
                lambda error=str(e):
                self.export_error(error)
            )

    # ------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------

    def export_csv(
        self,
        queries,
        names,
        filename
    ):
        """
        Write one CSV file per query.

        A single query keeps the exact file name chosen
        by the user, several queries get a suffix.

        Returns the list of (label, target) written and the
        error that stopped the run, if any.
        """

        base, extension = os.path.splitext(filename)

        total = len(queries)

        completed = []

        error_message = None

        used_files = set()

        for position, query in enumerate(
            queries,
            start=1
        ):

            name = names[position - 1]

            if total == 1:

                target = filename

            else:

                target = (
                    f"{base}_{position:02d}_"
                    f"{clean_file_name(name)}"
                    f"{extension}"
                )

                counter = 2

                while target.lower() in used_files:

                    target = (
                        f"{base}_{position:02d}_"
                        f"{clean_file_name(name)}"
                        f"_{counter}{extension}"
                    )

                    counter += 1

            used_files.add(target.lower())

            self.set_status(
                f"Query {position}/{total} "
                f"({name}): running..."
            )

            try:

                response = self.execute_query(
                    query,
                    "CSVWithNames"
                )

                self.write_csv_file(
                    response,
                    target,
                    name,
                    position,
                    total
                )

            except Exception as e:

                error_message = (
                    f"Query {position} \"{name}\" "
                    f"failed:\n\n{e}"
                )

                break

            completed.append(
                (name, os.path.basename(target))
            )

        if not completed:

            raise Exception(
                error_message
                or "No query produced a result."
            )

        return completed, error_message

    def write_csv_file(
        self,
        response,
        target,
        name,
        position,
        total
    ):

        total_bytes = 0

        with open(target, "wb") as file:

            while True:

                chunk = response.read(
                    1024 * 1024
                )

                if not chunk:
                    break

                file.write(chunk)

                total_bytes += len(chunk)

                mb = total_bytes / 1024 / 1024

                self.set_status(
                    f"Query {position}/{total} "
                    f"({name}): downloading "
                    f"{mb:.1f} MB"
                )

        response.close()

    # ------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------

    def export_excel(
        self,
        queries,
        names,
        filename
    ):
        """
        Write one workbook with one sheet per query.

        Returns the list of (sheet, rows) written and the
        error that stopped the run, if any.
        """

        try:
            from openpyxl import Workbook

        except ImportError:

            raise Exception(
                "Excel export requires openpyxl.\n\n"
                "Install it with:\n"
                "pip install openpyxl"
            )

        workbook = Workbook(
            write_only=True
        )

        total = len(queries)

        used_names = set()

        completed = []

        error_message = None

        for position, query in enumerate(
            queries,
            start=1
        ):

            name = clean_sheet_name(
                names[position - 1],
                used_names
            )

            self.set_status(
                f"Query {position}/{total} "
                f"({name}): running..."
            )

            try:

                response = self.execute_query(
                    query,
                    "CSVWithNames"
                )

            except Exception as e:

                error_message = (
                    f"Query {position} \"{name}\" "
                    f"failed:\n\n{e}"
                )

                break

            try:

                rows = self.write_worksheet(
                    workbook,
                    name,
                    used_names,
                    response,
                    position,
                    total
                )

            except Exception as e:

                # The sheet exists already and holds the
                # rows received before the failure
                error_message = (
                    f"Query {position} \"{name}\" failed "
                    f"while reading its result:\n\n{e}"
                )

                completed.append(
                    (name, "incomplete result")
                )

                break

            completed.append(
                (name, format_rows(rows))
            )

        if not completed:

            raise Exception(
                error_message
                or "No query produced a result."
            )

        self.set_status(
            "Saving the Excel file..."
        )

        workbook.save(filename)

        return completed, error_message

    def write_worksheet(
        self,
        workbook,
        name,
        used_names,
        response,
        position,
        total
    ):
        """
        Stream one query result into one sheet.

        Results longer than the Excel row limit continue
        on an extra sheet that repeats the header.
        """

        worksheet = workbook.create_sheet(name)

        text_stream = io.TextIOWrapper(
            response,
            encoding="utf-8",
            newline=""
        )

        reader = csv.reader(
            text_stream
        )

        header = None

        sheet_rows = 0

        total_rows = 0

        part = 1

        for row in reader:

            if header is None:
                header = row

            if sheet_rows >= EXCEL_ROW_LIMIT:

                part += 1

                worksheet = workbook.create_sheet(
                    clean_sheet_name(
                        f"{name} ({part})",
                        used_names
                    )
                )

                worksheet.append(header)

                sheet_rows = 1

            worksheet.append(row)

            sheet_rows += 1

            total_rows += 1

            if total_rows % 10000 == 0:

                self.set_status(
                    f"Query {position}/{total} "
                    f"({name}): writing "
                    f"{total_rows:,} rows"
                )

        text_stream.close()

        # The first row holds the column names
        return max(total_rows - 1, 0)

    # ------------------------------------------------------------
    # Results
    # ------------------------------------------------------------

    def export_success(
        self,
        filename,
        completed,
        error
    ):

        self.progress.stop()

        self.export_button.config(
            state="normal"
        )

        self.test_button.config(
            state="normal"
        )

        details = "\n".join(
            f"{position}. {label} - {value}"
            for position, (label, value) in enumerate(
                completed,
                start=1
            )
        )

        if error:

            self.status_var.set(
                f"Export incomplete: "
                f"{len(completed)} query(ies) saved"
            )

            messagebox.showwarning(
                "Export incomplete",
                f"{error}\n\n"
                f"The queries below were saved to:\n\n"
                f"{filename}\n\n"
                f"{details}"
            )

            return

        self.status_var.set(
            f"Export completed "
            f"({len(completed)} query(ies)): {filename}"
        )

        messagebox.showinfo(
            "Export completed",
            f"File successfully created:\n\n"
            f"{filename}\n\n"
            f"{details}"
        )

    def export_error(
        self,
        error
    ):

        self.progress.stop()

        self.export_button.config(
            state="normal"
        )

        self.test_button.config(
            state="normal"
        )

        self.status_var.set(
            "Export failed"
        )

        messagebox.showerror(
            "Error",
            error
        )


def main():

    root = tk.Tk()

    app = ClickHouseExporter(root)

    root.mainloop()


if __name__ == "__main__":
    main()
